import pandas as pd
import datetime
import xls_helpers
import openpyxl
import tkinter
from tkinter import filedialog
import os
import copy

min_date = datetime.datetime(2022,1,1)
max_date = datetime.datetime(2023,4,15)

print("Please select the rent roll file that you want to reformat.\n")
root = tkinter.Tk()
root.update()
file_path = filedialog.askopenfilename()
raw_rentroll = file_path
xls_helpers.mylogging('Loading ' + raw_rentroll + '...', 1)
wb_rentroll = openpyxl.load_workbook(raw_rentroll, data_only=True)

base_path = "/Users/mikephillips/My Drive/Cottonwood/C. Equity/1. Project Documents/Financials/Rent Rolls/Alyx/"

column_defs = {"Bldg-Unit": [1, 10],
               "Unit Type": [2, 10],
               "SQFT": [3, 7],
               "Resident Code": [4, 10],
               "Resident": [5, 25],
               "Market Rent": [6, 10],
               "Scheduled Rent": [7, 10],
               "Resident Deposit": [8, 5],
               "Other Deposit": [9, 5],
               "Lease Start": [10, 10],
               "Lease Expiration": [11, 10],
               "Move-Out": [12, 10],
               "Balance": [13, 10]}

for sheet in wb_rentroll.sheetnames:
    response = input("Do you want to reformat " + sheet + "?")
    if response == "Y":
        wb_rentroll_copy = copy.deepcopy(wb_rentroll)

        for s in wb_rentroll_copy.sheetnames:
            if s != sheet:
                del wb_rentroll_copy[s]

        for merge in list(wb_rentroll_copy[sheet].merged_cells):
            wb_rentroll_copy[sheet].unmerge_cells(str(merge))

        wb_rentroll_copy[sheet].delete_rows(1, 6)

        for each in column_defs.keys():
            wb_rentroll_copy[sheet].cell(row=1, column=column_defs[each][0]).value = each

        print(str(len(wb_rentroll_copy._external_links)) + " external links")

        items = wb_rentroll_copy._external_links
        for index, item in enumerate(items):
            Mystr = wb_rentroll_copy._external_links[index].file_link.Target
            print(Mystr)
            wb_rentroll_copy._external_links[index].file_link.Target = ''

        scrolling = True
        current_row = 2

        while scrolling:
            # print("Current row: " + str(current_row))
            if (wb_rentroll_copy[sheet].cell(row=current_row, column=2).value is None and
                    wb_rentroll_copy[sheet].cell(row=current_row, column=3).value is None and
                    wb_rentroll_copy[sheet].cell(row=current_row, column=4).value is None):
                # print("Deleting empty row")
                if wb_rentroll_copy[sheet].cell(row=current_row, column=1).value and \
                        ("Summary Groups" in wb_rentroll_copy[sheet].cell(row=current_row, column=1).value or
                         "Future Residents/Applicants" in wb_rentroll_copy[sheet].cell(row=current_row, column=1).value):
                    last_row = len(wb_rentroll_copy[sheet][1])
                    wb_rentroll_copy[sheet].delete_rows(current_row, current_row + last_row)
                    scrolling = False

                else:
                    wb_rentroll_copy[sheet].delete_rows(current_row)
            else:
                for i in range (10, 13):
                    # print("column  " + str(i))
                    # print(wb_rentroll_copy[sheet].cell(row=current_row, column=i).value)
                    # print(type(wb_rentroll_copy[sheet].cell(row=current_row, column=i).value))
                    if wb_rentroll_copy[sheet].cell(row=current_row, column=i).value:
                        if isinstance(wb_rentroll_copy[sheet].cell(row=current_row, column=i).value, datetime.datetime):
                            wb_rentroll_copy[sheet].cell(row=current_row, column=i).value = \
                                wb_rentroll_copy[sheet].cell(row=current_row, column=i).value.strftime("%m/%d/%Y")
                current_row += 1

        path, filename = os.path.split(file_path)
        name, ext = os.path.splitext(filename)
        wb_rentroll_copy.save(base_path + name + "_" + sheet + "_reformatted" + ext)

    else:
        print("Skipping " + sheet + "....")