import pandas as pd
from pandas import ExcelWriter
import xls_helpers
import openpyxl
import tkinter
from tkinter import filedialog
import os

base_path = "/Users/mikephillips/My Drive/Cottonwood/C. Equity/1. Project Documents/Financials/Rent Rolls/Raider/"

print("Please select the rent roll that you want to re-format.")
root = tkinter.Tk()
root.update()
raw_rentroll = filedialog.askopenfilename()

xls_helpers.mylogging('Loading ' + raw_rentroll + '...', 1)
wb_rentroll = openpyxl.load_workbook(raw_rentroll, data_only=True)

for merge in list(wb_rentroll["Summit On Nellis"].merged_cells):
    wb_rentroll["Summit On Nellis"].unmerge_cells(str(merge))

wb_rentroll["Summit On Nellis"].delete_rows(1, 6)

getting_cols = True
col = 1
col_dict = {}
while getting_cols:
    if wb_rentroll["Summit On Nellis"].cell(row=1, column=col).value is None:
        getting_cols = False
    else:
        col_dict[wb_rentroll["Summit On Nellis"].cell(row=1, column=col).value] = col
    col += 1

master_data = []

scrolling = True
currow = 2
while scrolling:
    print("Current row: " + str(currow))

    if wb_rentroll["Summit On Nellis"].cell(row=currow, column=1).value == "Status Summary":
        scrolling = False

    elif (wb_rentroll["Summit On Nellis"].cell(row=currow, column=col_dict["Bldg-Unit"]).value is not None and
        wb_rentroll["Summit On Nellis"].cell(row=currow, column=col_dict["Unit Status"]).value is not None):

        row_dict = {}
        for col_name in col_dict.keys():
            if col_name not in ["Charge Code", "Scheduled Charges"]:
                row_dict[col_name] = \
                    wb_rentroll["Summit On Nellis"].cell(row=currow, column=col_dict[col_name]).value
        sub_scrolling = True
        rrow = 0
        while sub_scrolling:
            if "Charge Code" in col_dict.keys():
                charge_code = \
                    wb_rentroll["Summit On Nellis"].cell(row=currow + rrow, column=col_dict["Charge Code"]).value
                if charge_code:
                    row_dict[charge_code] = \
                        wb_rentroll["Summit On Nellis"].cell(row=currow + rrow, column=col_dict["Scheduled Charges"]).value
                else:
                    sub_scrolling = False
            else:
                sub_scrolling = False
            rrow += 1
        master_data.append(row_dict)
        print("Unit " + str(row_dict["Bldg-Unit"]) + " added")

    currow += 1

master_data_df = pd.DataFrame.from_dict(master_data)

path, filename = os.path.split(raw_rentroll)
name, ext = os.path.splitext(filename)
writer = ExcelWriter(base_path + name + "_" + "_reformatted" + ext)
master_data_df.to_excel(writer, "Sheet1", index=False)
writer.save()
