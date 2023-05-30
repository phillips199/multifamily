from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import xls_helpers
import openpyxl
import tkinter
from tkinter import filedialog
import os

print("Please select the 'All Units' file that you want to reformat.")
root = tkinter.Tk()
root.update()
raw_rentroll = filedialog.askopenfilename()

path, filename = os.path.split(raw_rentroll)
name, ext = os.path.splitext(filename)
export_file = '/Users/mikephillips/My Drive/Cottonwood/C. Equity/1. Project Documents/Financials/Rent Rolls/Quinn/' + name + '_unmerged' + ext

xls_helpers.mylogging('Loading ' + raw_rentroll + '...', 1)
wb_rentroll = openpyxl.load_workbook(raw_rentroll, data_only=True)

for merge in list(wb_rentroll["Sheet1"].merged_cells):
    wb_rentroll["Sheet1"].unmerge_cells(str(merge))

wb_rentroll["Sheet1"].delete_rows(1, 16)

column_defs = {"Bldg-Unit": [1,10],
             "Unit Type": [6,10],
             "SQFT": [15,7],
             "Market Rent": [21,10],
             "Market Rent/SF": [28,12],
             "Scheduled Rent": [33,10],
             "Scheduled Rent/SF": [40,12],
             "Resident": [44,25],
             "Move-In": [61,10],
             "Lease Start": [67,10],
             "Lease End": [73,10],
             "Deposit Held": [81,10],
             "Made Ready": [87,10]}

for each in column_defs.keys():

    wb_rentroll["Sheet1"].cell(row=1, column=column_defs[each][0]).value = each

wb_rentroll.save("Unmerged_Rentroll.xlsx")

scrolling = True
current_row = 1
while scrolling:
    print("Current row: " + str(current_row))
    if (wb_rentroll["Sheet1"].cell(row=current_row, column=6).value is None and
        wb_rentroll["Sheet1"].cell(row=current_row, column=15).value is None and
        wb_rentroll["Sheet1"].cell(row=current_row, column=21).value is None):
        if wb_rentroll["Sheet1"].cell(row=current_row, column=1).value and "total for property" in wb_rentroll["Sheet1"].cell(row=current_row, column=1).value:
            last_row = len(wb_rentroll["Sheet1"][1])
            wb_rentroll["Sheet1"].delete_rows(current_row,current_row + last_row)
            scrolling = False
        else:
            print("Deleting empty row")
            wb_rentroll["Sheet1"].delete_rows(current_row)
    else:
        print("Skipping unit " + wb_rentroll["Sheet1"].cell(row=current_row, column=1).value)
        wb_rentroll["Sheet1"].row_dimensions[current_row].height = 13
        current_row += 1

scrolling = True
current_col = 1
while scrolling:
    if wb_rentroll["Sheet1"].cell(row=1, column=current_col).value is None:
        wb_rentroll["Sheet1"].delete_cols(current_col)
    else:
        column_name = wb_rentroll["Sheet1"].cell(row=1, column=current_col).value
        column_ltr = get_column_letter(current_col)
        wb_rentroll["Sheet1"].column_dimensions[column_ltr].width = column_defs[column_name][1]
        wb_rentroll["Sheet1"].column_dimensions[column_ltr].font = Font(size=11,name="Arial")
        print("Skipping " + column_name)
        if column_name == "Made Ready":
            scrolling = False
        else:
            current_col += 1

wb_rentroll.save(export_file)