from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import xls_helpers
import datetime
import openpyxl
import tkinter
from tkinter import filedialog
import os
import copy

print("Please select the rent roll file that you want to reformat.\n")
root = tkinter.Tk()
root.update()
file_path = filedialog.askopenfilename()
raw_rentroll = file_path
xls_helpers.mylogging('Loading ' + raw_rentroll + '...', 1)
wb_rentroll = openpyxl.load_workbook(raw_rentroll, data_only=True)

base_path = "/Users/mikephillips/My Drive/Cottonwood/C. Equity/1. Project Documents/Financials/Rent Rolls/"

old_file_format = input("Is this an old file format?")
if old_file_format == "Y":

    for merge in list(wb_rentroll["Sheet"].merged_cells):
        wb_rentroll["Sheet"].unmerge_cells(str(merge))

    wb_rentroll["Sheet"].delete_rows(1, 6)
    wb_rentroll.save("Unmerged_Rentroll.xlsx")

    column_defs = {"Bldg-Unit": [1,10],
                 "Unit Type": [2,10],
                 "SQFT": [4,7],
                 "Resident": [5,25],
                 "Status": [9,5],
                 "Market Rent": [11,10],
                 "Scheduled Rent": [15,10],
                 "Other": [18,12],
                 "Credits": [20,12],
                 "Total": [22,10],
                 "Move-In": [25,10],
                 "Lease Start": [27,10],
                 "Lease End": [29,10],
                 "Expected Move-Out": [30,10],
                 "Surety Bonds": [32,10],
                 "Deposits": [35,10],
                 "Balance": [36,10]}

    for each in column_defs.keys():

        wb_rentroll["Sheet"].cell(row=1, column=column_defs[each][0]).value = each

    wb_rentroll.save("Unmerged_Rentroll.xlsx")

    scrolling = True
    current_row = 1
    while scrolling:
        print("Current row: " + str(current_row))
        if (wb_rentroll["Sheet"].cell(row=current_row, column=2).value is None and
            wb_rentroll["Sheet"].cell(row=current_row, column=3).value is None and
            wb_rentroll["Sheet"].cell(row=current_row, column=4).value is None):
            if wb_rentroll["Sheet"].cell(row=current_row, column=1).value and "Property Occupancy" in \
                wb_rentroll["Sheet"].cell(row=current_row, column=1).value:
                last_row = len(wb_rentroll["Sheet"][1])
                wb_rentroll["Sheet"].delete_rows(current_row,current_row + last_row)
                scrolling = False
            else:
                # print("Deleting empty row")
                wb_rentroll["Sheet"].delete_rows(current_row)
        else:
            print("Skipping unit " + wb_rentroll["Sheet"].cell(row=current_row, column=1).value)
            wb_rentroll["Sheet"].row_dimensions[current_row].height = 13
            current_row += 1

    scrolling = True
    current_col = 1
    while scrolling:
        if wb_rentroll["Sheet"].cell(row=1, column=current_col).value is None:
            wb_rentroll["Sheet"].delete_cols(current_col)
        else:
            column_name = wb_rentroll["Sheet"].cell(row=1, column=current_col).value
            column_ltr = get_column_letter(current_col)
            wb_rentroll["Sheet"].column_dimensions[column_ltr].width = column_defs[column_name][1]
            wb_rentroll["Sheet"].column_dimensions[column_ltr].font = Font(size=11,name="Arial")
            # print("Skipping " + column_name)
            if column_name == "Balance":
                scrolling = False
            else:
                current_col += 1

    path, filename = os.path.split(file_path)
    name, ext = os.path.splitext(filename)
    wb_rentroll.save(base_path + name + "_reformatted" + ext)

else:

    # wb_rentroll["Sheet"].delete_cols(15,50)
    for sheet in wb_rentroll.sheetnames:
        response = input("Do you want to reformat " + sheet + "?")
        if response == "Y":
            wb_rentroll_copy = copy.deepcopy(wb_rentroll)
            for s in wb_rentroll_copy.sheetnames:
                if s != sheet:
                    del wb_rentroll_copy[s]

            month = wb_rentroll_copy[sheet].cell(row=4, column=1).value
            wb_rentroll_copy[sheet].delete_rows(1, 6)
            wb_rentroll_copy[sheet].delete_cols(15, 90)

            scrolling = True
            current_row = 1
            while scrolling:
                # print("Current row: " + str(current_row))
                if (wb_rentroll_copy[sheet].cell(row=current_row, column=4).value is None and
                    wb_rentroll_copy[sheet].cell(row=current_row, column=5).value is None and
                    wb_rentroll_copy[sheet].cell(row=current_row, column=6).value is None):
                    # print("Deleting empty row")
                    wb_rentroll_copy[sheet].delete_rows(current_row)
                else:
                    if wb_rentroll_copy[sheet].cell(row=current_row, column=1).value and "Status Summary" in \
                            wb_rentroll_copy[sheet].cell(row=current_row, column=1).value:
                        last_row = len(wb_rentroll_copy[sheet][1])
                        wb_rentroll_copy[sheet].delete_rows(current_row, current_row + last_row)
                        scrolling = False

                    else:
                        wb_rentroll_copy[sheet].cell(row=current_row, column=1).value = \
                            str(wb_rentroll_copy[sheet].cell(row=current_row, column=1).value)
                        for i in range(6, 10):
                            if wb_rentroll_copy[sheet].cell(row=current_row, column=i).value:
                                if isinstance(wb_rentroll_copy[sheet].cell(row=current_row, column=i).value,
                                              datetime.datetime):
                                    wb_rentroll_copy[sheet].cell(row=current_row, column=i).value = \
                                        wb_rentroll_copy[sheet].cell(row=current_row, column=i).value.strftime(
                                            "%m/%d/%Y")

                        current_row += 1

            path, filename = os.path.split(file_path)
            name, ext = os.path.splitext(filename)
            wb_rentroll_copy.save(base_path + name + "_" + sheet + "_reformatted" + ext)

        else:
            print("Skipping " + sheet + "....")