import xls_helpers
import openpyxl
import tkinter
from tkinter import filedialog
import os

print("Please select the tracker file that you want to reformat.")
raw_tracker = filedialog.askopenfilename()
print(raw_tracker)

xls_helpers.mylogging('Loading ' + raw_tracker + '...', 1)
wb_tracker = openpyxl.load_workbook(raw_tracker, data_only=True)

for sheet in wb_tracker.sheetnames:
    if sheet != "Tracker":
        del wb_tracker[sheet]
        print(sheet + " deleted!")

print("About to unmerge call cells...")
for merge in list(wb_tracker["Tracker"].merged_cells):
    wb_tracker["Tracker"].unmerge_cells(str(merge))

print("About to delete rows...")
wb_tracker["Tracker"].delete_rows(1, 13)
print("About to delete column 1...")
wb_tracker["Tracker"].delete_cols(1)
print("About to remove autofilter...")
wb_tracker["Tracker"].auto_filter.ref = None
print("About to remove freeze panes...")
wb_tracker["Tracker"].freeze_panes = None
print("About to set row heights...")
wb_tracker["Tracker"].row_dimensions[1].height = 52
for i in range(2, wb_tracker["Tracker"].max_row):
    wb_tracker["Tracker"].row_dimensions[i].height = 13
print("About to ungroup rows...")
wb_tracker["Tracker"].row_dimensions.group(1, wb_tracker["Tracker"].max_row + 1, outline_level=0)
print("About to delete images...")
wb_tracker["Tracker"]._images = []
print("About to save workbbook...")

path, filename = os.path.split(raw_tracker)
name, ext = os.path.splitext(filename)
export_file = '/Users/mikephillips/My Drive/Cottonwood/C. Equity/1. Project Documents/Financials/Rent Rolls/Quinn/' + name + '_unmerged' + ext

wb_tracker.save(export_file)
